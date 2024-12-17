from django.shortcuts import render
from django.shortcuts import render
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from io import BytesIO
import xlsxwriter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from order.models import Order,OrderItem
from django.db.models import Sum
from cart.models import Cart
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import timedelta
from datetime import datetime
from django.views.decorators.cache import never_cache 
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from datetime import datetime
from products.models import Product,ProductImage,ProductSize
from django.conf import settings
from xhtml2pdf import pisa


# def admin_required(function):
#     return user_passes_test(
#         lambda user: user.is_superuser,
#         login_url='misc_pages:custom_404')(function)
 
 
@never_cache
# @admin_required
def sales_report(request):
    start_date = request.GET.get('start_date')
    
    end_date = request.GET.get('end_date')
    report_type = request.GET.get('report_type', 'daily')
     

    orders = Order.objects.all()
    
    if report_type == 'daily':
        orders = orders.filter(created_at__date=timezone.now().date())
    elif report_type == 'weekly':
        start_of_week = timezone.now().date() - timedelta(days=7)
        orders = orders.filter(created_at__date__gte=start_of_week)
    elif report_type == 'monthly':
        current_month = timezone.now().month
        current_year = timezone.now().year
        orders = orders.filter(created_at__month=current_month, created_at__year=current_year)
    if report_type == 'custom'and start_date and end_date:
           start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
           end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
           orders = orders.filter(created_at__range=[start_date, end_date])
 
    total_sales_count = orders.count()
    total_sales_amount = orders.aggregate(Sum('total_price'))['total_price__sum']
    total_discount = Cart.objects.filter(user=request.user).aggregate(Sum('discount'))['discount__sum']

    context = {
        'orders': orders,
        'total_sales_count': total_sales_count,
        'total_sales_amount': total_sales_amount,
        'total_discount': total_discount,
        'report_type':report_type,
        'end_date':end_date,
        'start_date':start_date
    }

    return render(request, 'reports/report.html', context)



def download_report(request, report_format):
    report_type = request.GET.get('report_type', 'daily')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date:
        start_date = datetime.strptime(start_date, "%b. %d, %Y").strftime("%Y-%m-%d")
    if end_date:
        end_date = datetime.strptime(end_date, "%b. %d, %Y").strftime("%Y-%m-%d")

  
 
    orders = Order.objects.all()

    if report_type == 'daily':
        orders = orders.filter(created_at__date=timezone.now().date())
    elif report_type == 'weekly':
        start_of_week = timezone.now().date() - timedelta(days=7)
        orders = orders.filter(created_at__date__gte=start_of_week)
    elif report_type == 'monthly':
        current_month = timezone.now().month
        current_year = timezone.now().year
        orders = orders.filter(created_at__month=current_month, created_at__year=current_year)
    elif report_type == 'custom' and start_date and end_date:
        orders = orders.filter(created_at__range=[start_date, end_date])

   
    if report_format == 'pdf':
        return generate_sales_pdf(orders)
    elif report_format == 'excel':
        return generate_sales_excel(orders)
    

def generate_sales_excel(orders):
   
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

   
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    headers = ["Order ID", "Customer", "Total Amount", "Discount", "Payment Type", "Order Date"]
         
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

     
    row = 2
    for order in orders:
        order_date = order.created_at.replace(tzinfo=None)
        cart = Cart.objects.get(user=order.user)  
        discount = cart.discount
        ws.cell(row=row, column=1, value=order.order_id)
        ws.cell(row=row, column=2, value=order.user.username)
        ws.cell(row=row, column=3, value=order.total_price)
        ws.cell(row=row, column=4, value=discount)
        ws.cell(row=row, column=5, value=order.payment_type)
        ws.cell(row=row, column=6, value=order_date.strftime('%Y-%m-%d %H:%M:%S'))   
        
        row += 1

    
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    response = HttpResponse(file_stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'

    return response




# def generate_sales_pdf(orders):
  
#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="sales_report.pdf"'

#     buffer = BytesIO()
#     p = canvas.Canvas(buffer, pagesize=letter)
    
   
#     margin_left = 30
#     margin_top = 750
#     column_width = 100   

    
#     p.setFont('Helvetica-Bold', 18)
#     p.setFillColor(colors.white)
#     p.rect(0, 740, 600, 40)   
#     p.setFillColor(colors.blue)
#     p.drawString(180, 755, f"Sales Report: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    
#     p.setFont('Helvetica-Bold', 12)
#     p.setFillColor(colors.white)
#     p.rect(margin_left, margin_top - 50, 550, 20, fill=1)  
#     p.setFillColor(colors.black)
#     p.drawString(margin_left + 10, margin_top - 40, "Order ID")
#     p.drawString(margin_left + column_width, margin_top - 40, "Total Price")
#     p.drawString(margin_left + 2 * column_width, margin_top - 40, "Discount")
#     p.drawString(margin_left + 3 * column_width, margin_top - 40, "Payment Type")
#     p.drawString(margin_left + 4 * column_width, margin_top - 40, "Order Date")

    
#     y_position = margin_top - 70
#     p.setFont('Helvetica', 10)

    
#     for order in orders:
#         cart = Cart.objects.get(user=order.user)
#         discount = cart.discount
#         row_color = colors.grey if y_position % 2 == 0 else colors.whitesmoke
#         p.setFillColor(row_color)
#         p.rect(margin_left, y_position - 10, 550, 20, fill=1)  
#         p.setFillColor(colors.black)

#         p.drawString(margin_left + 10, y_position, str(order.order_id))
#         p.drawString(margin_left + column_width + 10, y_position, f"₹{order.total_price}")
#         p.drawString(margin_left + 2 * column_width + 10, y_position, f"₹{discount}")
#         p.drawString(margin_left + 3 * column_width + 10, y_position, order.payment_type)
#         p.drawString(margin_left + 4 * column_width + 10, y_position, order.created_at.strftime('%Y-%m-%d %H:%M:%S'))

#         y_position -= 20

        
#         if y_position < 60:
#             p.showPage() 
#             p.setFont('Helvetica-Bold', 18)
#             p.setFillColor(colors.blue)
#             p.drawString(180, 755, f"Sales Report: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}")
#             p.setFont('Helvetica-Bold', 12)
#             p.setFillColor(colors.white)
#             p.rect(margin_left, margin_top - 50, 550, 20, fill=1)
#             p.setFillColor(colors.black)
#             p.drawString(margin_left + 10, margin_top - 40, "Order ID")
#             p.drawString(margin_left + column_width, margin_top - 40, "Total Price")
#             p.drawString(margin_left + 2 * column_width, margin_top - 40, "Discount")
#             p.drawString(margin_left + 3 * column_width, margin_top - 40, "Payment Type")
#             p.drawString(margin_left + 4 * column_width, margin_top - 40, "Order Date")
#             y_position = margin_top - 70   

    
#     p.setFont('Helvetica', 8)
#     p.drawString(270, 30, f"Page {p.getPageNumber()}")
#     p.showPage()
#     p.save()
#     buffer.seek(0)
#     response.write(buffer.read())
#     return response


def generate_sales_pdf(orders):
    total_sales_value = sum(order.total_price for order in orders)
    average_order_value = total_sales_value / len(orders) if orders else 0
    discount=total_price-order.total_price
    context = {
        'orders': orders,
        'timestamp': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
        'total_sales_value': total_sales_value,
        'average_order_value': average_order_value,
    }

    html_content = render_to_string('reports/sales.html', context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="sales_report.pdf"'

    pisa_status = pisa.CreatePDF(html_content, dest=response)

    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)

    return response


 
def admin_dashboard(request):
    filter_option = request.GET.get('filter', 'monthly')
    if filter_option == 'monthly':
        orders = Order.objects.filter(created_at__month=datetime.now().month)
    elif filter_option == 'yearly':
        orders = Order.objects.filter(created_at__year=datetime.now().year)
    elif filter_option == 'daily':
        orders = Order.objects.filter(created_at__day=datetime.now().day)
    else:
        orders = Order.objects.all()

    total_sales = orders.aggregate(Sum('total_price'))['total_price__sum'] or 0

    best_selling_products = OrderItem.objects.values('product__id', 'product__name', 'product__category__id', 'product__category__image') \
        .annotate(total_sales=Sum('quantity')) \
        .order_by('-total_sales')[:10]
    for product in best_selling_products:
        try:
            product_obj = Product.objects.get(id=product['product__id'])
            product['image_url'] = product_obj.additional_images.first().image.url if product_obj.additional_images.exists() else '/media/default-product.jpg'
        except Product.DoesNotExist:
            product['image_url'] = '/media/default-product.jpg'

    
    best_selling_categories = OrderItem.objects.values(
        'product__Types__category',  
        'product__Types__id',         
        'product__Types__image',       
        'product__category__brand_name',   
        'product__category__image'  
    ).annotate(total_sales=Sum('quantity')).order_by('-total_sales')

   
    for category in best_selling_categories:
       
        category_image = category['product__Types__image']
        if category_image:
            category['category_image_url'] = settings.MEDIA_URL + category_image
        else:
            category['category_image_url'] = '/media/default-category.jpg'

        
        brand_image = category['product__category__image']
        
        if brand_image:
            category['brand_image_url'] = settings.MEDIA_URL + brand_image
        else:
            category['brand_image_url'] = '/media/default-brand.jpg'
        

    
    best_selling_brands = OrderItem.objects.values('product__category__brand_name') \
        .annotate(total_sales=Sum('quantity')) \
        .order_by('-total_sales')[:10]

     
    best_selling_brands_names = [brand['product__category__brand_name'] for brand in best_selling_brands]
    best_selling_brands_sales = [brand['total_sales'] for brand in best_selling_brands]

    sales_by_month = [0] * 12  
    if filter_option == 'monthly' or filter_option == 'yearly':
        orders_by_month = orders.values('created_at__month') \
            .annotate(total_sales=Sum('total_price')) \
            .order_by('created_at__month')
        for sale in orders_by_month:
            month_index = sale['created_at__month'] - 1  
            sales_by_month[month_index] = float(sale['total_sales'])   
    elif filter_option == 'daily':
        orders_by_day = orders.values('created_at__day') \
            .annotate(total_sales=Sum('total_price')) \
            .order_by('created_at__day')
        sales_by_month = [0] * 31  
        for sale in orders_by_day:
            day_index = sale['created_at__day'] - 1    
            sales_by_month[day_index] = float(sale['total_sales'])   

    context = {
        'total_sales': total_sales,
        'orders': orders,
        'filter_option': filter_option,
        'best_selling_products': best_selling_products,
        'best_selling_categories': best_selling_categories,
        'best_selling_brands': best_selling_brands,
        'best_selling_brands_names': best_selling_brands_names,   
        'best_selling_brands_sales': best_selling_brands_sales,  
        'sales_by_month': sales_by_month,  
    }

    return render(request, 'reports/dashboard.html', context)
